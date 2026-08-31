import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from automation.tablas_normativas.engine import FIELDS, audit_table, load_rule_catalog, process_file


class TablasNormativasEngineTests(unittest.TestCase):
    def base_row(self):
        row = {field: "" for field in FIELDS}
        row.update({
            "COMUNA": "PRUEBA",
            "RIALCOMSII": "99999",
            "CODIGO_PRC": "Z1",
            "ZONA": "Z1",
            "CONSTRUCCION": "0,8",
            "OCUPACION": "0,6",
            "PISOS_MAX": "4",
            "ALTURA_MIN": 3,
            "ALTURA_MAX": 12,
            "AGRUPAMIENTO": "pareado / aislado",
        })
        return row

    def test_preserva_exactamente_35_campos_y_normaliza_formato(self):
        row = self.base_row()
        result = audit_table(FIELDS.copy(), [row])
        self.assertEqual(list(result["rows"][0].keys()), FIELDS)
        self.assertEqual(result["rows"][0]["CONSTRUCCION"], 0.8)
        self.assertEqual(result["rows"][0]["PISOS_MAX"], 4)
        self.assertEqual(result["rows"][0]["AGRUPAMIENTO"], "AISLADO; PAREADO")
        self.assertEqual(result["input_rows"], 1)
        self.assertEqual(result["output_rows"], 1)
        self.assertTrue(any(f["status"] == "NORMALIZACIÓN DE FORMATO" for f in result["findings"]))

    def test_no_corrige_ocr_sin_fuente(self):
        row = self.base_row()
        row["CONSTRUCCION"] = "P.2"
        result = audit_table(FIELDS.copy(), [row])
        self.assertEqual(result["rows"][0]["CONSTRUCCION"], "P.2")
        self.assertTrue(any(f["field"] == "CONSTRUCCION" and f["status"] == "POSIBLE ERROR" for f in result["findings"]))

    def test_regla_fuente_especifica_puede_autocorregir_campo_no_codigo(self):
        row = self.base_row()
        row["COMUNA"] = "Puerto Octay"
        row["CONSTRUCCION"] = "P.2"
        catalog = {"exact_rules": [{
            "id": "test-octay",
            "comuna": "Puerto Octay",
            "field": "CONSTRUCCION",
            "original": "P.2",
            "corrected": 0.2,
            "confidence": "ALTA",
            "auto_apply": True,
            "source": "Ordenanza oficial",
            "page": 10,
            "reason": "Regla de prueba respaldada por fuente."
        }]}
        result = audit_table(FIELDS.copy(), [row], catalog)
        self.assertEqual(result["rows"][0]["CONSTRUCCION"], 0.2)
        finding = next(f for f in result["findings"] if f["rule_id"] == "test-octay")
        self.assertEqual(finding["source"], "Ordenanza oficial")
        self.assertEqual(finding["page"], "10")

    def test_codigo_prc_se_preserva_si_regla_no_autoriza_cambio(self):
        row = self.base_row()
        row["COMUNA"] = "PEÑALOLÉN"
        row["CODIGO_PRC"] = "15152-PARQUE METROPOLITANO"
        row["ZONA"] = "PARQUE METROPOLITANO"
        catalog = {"exact_rules": [{
            "id": "regla-no-autorizada",
            "comuna": "PEÑALOLÉN",
            "field": "CODIGO_PRC",
            "original": "15152-PARQUE METROPOLITANO",
            "corrected": "15152-5.2.2",
            "confidence": "ALTA",
            "auto_apply": True,
            "source": "PRMS",
            "reason": "Sólo homogeneización."
        }]}
        result = audit_table(FIELDS.copy(), [row], catalog)
        self.assertEqual(result["rows"][0]["CODIGO_PRC"], "15152-PARQUE METROPOLITANO")
        finding = next(f for f in result["findings"] if f["rule_id"] == "regla-no-autorizada")
        self.assertEqual(finding["status"], "POSIBLE ERROR")

    def test_codigo_prc_solo_cambia_con_autorizacion_explicita_y_contexto(self):
        row = self.base_row()
        row["COMUNA"] = "PEÑALOLÉN"
        row["CODIGO_PRC"] = "15152-SM-1"
        row["ZONA"] = "R11"
        catalog = {"exact_rules": [{
            "id": "test-r11-codigo-confirmado",
            "comuna": "PEÑALOLÉN",
            "field": "CODIGO_PRC",
            "original": "15152-SM-1",
            "corrected": "15152-R11",
            "where": {"ZONA": "R11"},
            "confidence": "ALTA",
            "auto_apply": True,
            "allow_codigo_prc_change": True,
            "source": "Decreto Alcaldicio N° 1200/3504",
            "page": "Art. 31",
            "reason": "La fila corresponde a la zona R11 y el código SM-1 es inconsistente."
        }]}
        result = audit_table(FIELDS.copy(), [row], catalog)
        self.assertEqual(result["rows"][0]["CODIGO_PRC"], "15152-R11")
        finding = next(f for f in result["findings"] if f["rule_id"] == "test-r11-codigo-confirmado")
        self.assertEqual(finding["status"], "ERROR CONFIRMADO")
        self.assertEqual(finding["confidence"], "ALTA")

    def test_regla_condicionada_no_se_aplica_fuera_de_zona(self):
        row = self.base_row()
        row["COMUNA"] = "PEÑALOLÉN"
        row["CODIGO_PRC"] = "15152-SM-1"
        row["ZONA"] = "R1"
        catalog = load_rule_catalog(Path("config/tablas_normativas_reglas.json"))
        result = audit_table(FIELDS.copy(), [row], catalog)
        self.assertEqual(result["rows"][0]["CODIGO_PRC"], "15152-SM-1")

    def test_filas_con_atributos_iguales_se_conservan(self):
        row1 = self.base_row()
        row2 = self.base_row().copy()
        result = audit_table(FIELDS.copy(), [row1, row2])
        self.assertEqual(result["input_rows"], 2)
        self.assertEqual(result["output_rows"], 2)
        self.assertEqual(len(result["rows"]), 2)
        duplicate_finding = next(f for f in result["findings"] if f["field"] == "FILA")
        self.assertEqual(duplicate_finding["status"], "POSIBLE ERROR")
        self.assertIn("no se fusionan ni eliminan", duplicate_finding["reason"])

    def test_process_file_genera_normalizada_qa_status_y_misma_cantidad_filas(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            csv_path = root / "PRC_PRUEBA_35_CAMPOS.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=FIELDS)
                writer.writeheader()
                writer.writerow(self.base_row())
                writer.writerow(self.base_row())
            result = process_file(csv_path, root / "normalizadas", root / "qa")
            self.assertTrue(Path(result["normalized_path"]).exists())
            self.assertTrue(Path(result["qa_path"]).exists())
            self.assertTrue(Path(result["status_path"]).exists())
            self.assertEqual(result["input_rows"], 2)
            self.assertEqual(result["output_rows"], 2)
            wb = load_workbook(result["normalized_path"], read_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            self.assertEqual(len(rows) - 1, 2)
            self.assertEqual(list(rows[0]), FIELDS)


if __name__ == "__main__":
    unittest.main()
