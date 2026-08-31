import json
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from automation.tablas_normativas.engine import FIELDS
from automation.tablas_normativas.runner import run


class TablasNormativasRunnerTests(unittest.TestCase):
    def _make_gpkg(self, path: Path, codes: list[str]):
        path.parent.mkdir(parents=True, exist_ok=True)
        connection = sqlite3.connect(path)
        try:
            connection.executescript(
                """
                CREATE TABLE gpkg_contents (table_name TEXT PRIMARY KEY, data_type TEXT, identifier TEXT);
                CREATE TABLE gpkg_geometry_columns (table_name TEXT, column_name TEXT);
                CREATE TABLE PRC_PRUEBA (
                  fid INTEGER PRIMARY KEY, geom BLOB, ZONA TEXT, CODIGO_PRC TEXT, RIALCOMSII INTEGER
                );
                INSERT INTO gpkg_contents VALUES ('PRC_PRUEBA', 'features', 'PRC_PRUEBA');
                INSERT INTO gpkg_geometry_columns VALUES ('PRC_PRUEBA', 'geom');
                """
            )
            for index, code in enumerate(codes, start=1):
                connection.execute(
                    "INSERT INTO PRC_PRUEBA(fid, geom, ZONA, CODIGO_PRC, RIALCOMSII) VALUES (?, X'00', ?, ?, 99999)",
                    (index, code, code),
                )
            connection.commit()
        finally:
            connection.close()

    def _make_master(self, path: Path, codes: list[str]):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "PRC_PRUEBA"
        sheet.append(FIELDS)
        for index, code in enumerate(codes, start=1):
            row = {field: "" for field in FIELDS}
            row.update({
                "COMUNA": "PRUEBA",
                "RIALCOMSII": 99999,
                "CODIGO_PRC": code,
                "ZONA": code,
                "SUB_PREDIAL": 100 + index,
                "CONSTRUCCION": "0,8",
                "OCUPACION": "0,6",
            })
            sheet.append([row[field] for field in FIELDS])
        workbook.save(path)

    def _json(self, path: Path, content):
        path.write_text(json.dumps(content, ensure_ascii=False), encoding="utf-8")

    def _fixture(self, root: Path, polygon_codes, table_codes, coverage="COMPLETA"):
        prc_root = root / "IPT_Metropolitana" / "PRC" / "PRUEBA" / "PRC Trabajado"
        gpkg = prc_root / "PRC_PRUEBA.gpkg"
        self._make_gpkg(gpkg, polygon_codes)
        master = root / "PRC_SQL2.xlsx"
        self._make_master(master, table_codes)
        exact = root / "exact.json"
        conditional = root / "conditional.json"
        aliases = root / "aliases.json"
        coverage_path = root / "coverage.json"
        self._json(exact, {"exact_rules": []})
        self._json(conditional, {"conditional_rules": []})
        self._json(aliases, {"por_comuna": {}})
        self._json(coverage_path, {"por_comuna": {"PRUEBA": {"estado": coverage}}})
        return prc_root.parents[3], master, exact, conditional, aliases, coverage_path

    def test_publica_solo_si_vinculo_y_fuentes_estan_completos(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            prc_root, master, exact, conditional, aliases, coverage = self._fixture(
                root, ["Z1", "Z1", "Z2"], ["Z1", "Z1", "Z2"], "COMPLETA"
            )
            output = root / "NORMALIZADAS"
            state = root / "estado.json"
            result = run(
                prc_root=prc_root,
                master_path=master,
                output_dir=output,
                exact_rules_path=exact,
                conditional_rules_path=conditional,
                aliases_path=aliases,
                coverage_path=coverage,
                state_path=state,
            )
            item = result["comunas"]["PRUEBA"]
            self.assertEqual(item["estado"], "LISTA PARA STAGING")
            normalized = output / "PRC_PRUEBA_NORMALIZADO.xlsx"
            self.assertTrue(normalized.exists())
            workbook = load_workbook(normalized, read_only=True, data_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            self.assertEqual(len(rows) - 1, 3)
            self.assertEqual(list(rows[0]), FIELDS)

    def test_no_publica_si_fuentes_estan_incompletas(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            prc_root, master, exact, conditional, aliases, coverage = self._fixture(
                root, ["Z1"], ["Z1"], "PARCIAL"
            )
            output = root / "NORMALIZADAS"
            result = run(
                prc_root=prc_root,
                master_path=master,
                output_dir=output,
                exact_rules_path=exact,
                conditional_rules_path=conditional,
                aliases_path=aliases,
                coverage_path=coverage,
                state_path=root / "estado.json",
            )
            self.assertEqual(result["comunas"]["PRUEBA"]["estado"], "FUENTES INCOMPLETAS")
            self.assertFalse((output / "PRC_PRUEBA_NORMALIZADO.xlsx").exists())

    def test_no_publica_si_hay_codigo_sin_vinculo(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            prc_root, master, exact, conditional, aliases, coverage = self._fixture(
                root, ["Z1", "Z2"], ["Z1"], "COMPLETA"
            )
            output = root / "NORMALIZADAS"
            result = run(
                prc_root=prc_root,
                master_path=master,
                output_dir=output,
                exact_rules_path=exact,
                conditional_rules_path=conditional,
                aliases_path=aliases,
                coverage_path=coverage,
                state_path=root / "estado.json",
            )
            item = result["comunas"]["PRUEBA"]
            self.assertEqual(item["estado"], "ERROR VÍNCULO")
            self.assertIn("Z2", item["sin_normativa"])
            self.assertFalse((output / "PRC_PRUEBA_NORMALIZADO.xlsx").exists())


if __name__ == "__main__":
    unittest.main()
