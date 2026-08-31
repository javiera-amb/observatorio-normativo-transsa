from __future__ import annotations

import csv
import re
import sqlite3
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from . import engine as base


def _key(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


def _feature_tables(connection: sqlite3.Connection) -> list[str]:
    try:
        rows = connection.execute(
            "SELECT table_name FROM gpkg_contents WHERE data_type='features' ORDER BY table_name"
        ).fetchall()
    except sqlite3.DatabaseError as exc:
        raise ValueError("El archivo no es un GeoPackage válido o no contiene gpkg_contents.") from exc
    return [str(row[0]) for row in rows]


def _choose_feature_table(connection: sqlite3.Connection, preferred_layer: str | None = None) -> str:
    tables = _feature_tables(connection)
    if not tables:
        raise ValueError("El PRC no contiene ninguna capa vectorial de tipo features.")
    if preferred_layer:
        matches = [name for name in tables if _key(name) == _key(preferred_layer)]
        if len(matches) == 1:
            return matches[0]
        raise ValueError(f"No se encontró la capa PRC indicada: {preferred_layer}.")
    if len(tables) == 1:
        return tables[0]
    preferred = [
        name for name in tables
        if any(token in _key(name) for token in ("ZONIFIC", "PRC", "ZONA", "NORMAT"))
    ]
    if len(preferred) == 1:
        return preferred[0]
    raise ValueError(
        "El GeoPackage contiene varias capas de polígonos y no se puede determinar el PRC de forma inequívoca."
    )


def _column_name(connection: sqlite3.Connection, layer: str, wanted: str) -> str | None:
    quoted = layer.replace('"', '""')
    columns = [str(row[1]) for row in connection.execute(f'PRAGMA table_info("{quoted}")').fetchall()]
    matches = [column for column in columns if _key(column) == _key(wanted)]
    return matches[0] if len(matches) == 1 else None


def read_prc_codes(
    prc_path: str | Path,
    preferred_layer: str | None = None,
) -> tuple[int, str, list[str]]:
    path = Path(prc_path)
    if not path.exists():
        raise FileNotFoundError(path)
    if path.suffix.lower() != ".gpkg":
        raise ValueError("El PRC productivo debe entregarse como GeoPackage (.gpkg).")

    connection = sqlite3.connect(path)
    try:
        layer = _choose_feature_table(connection, preferred_layer)
        code_column = _column_name(connection, layer, "CODIGO_PRC")
        if not code_column:
            raise ValueError("La capa PRC no contiene un campo CODIGO_PRC inequívoco.")
        quoted_layer = layer.replace('"', '""')
        quoted_code = code_column.replace('"', '""')
        rows = connection.execute(
            f'SELECT "{quoted_code}" FROM "{quoted_layer}" ORDER BY rowid'
        ).fetchall()
        codes = [str(row[0] or "").strip() for row in rows]
        return len(rows), layer, codes
    finally:
        connection.close()


def _read_table(
    table_path: Path,
    table_sheet: str | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = table_path.suffix.lower()
    if suffix == ".csv":
        with table_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = [str(item or "").strip() for item in (reader.fieldnames or [])]
            rows = [{str(k or "").strip(): v for k, v in row.items()} for row in reader]
        return headers, rows
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(table_path, read_only=True, data_only=True)
        if table_sheet:
            if table_sheet not in workbook.sheetnames:
                raise ValueError(f"No existe la hoja {table_sheet} en la tabla normativa.")
            sheet = workbook[table_sheet]
        else:
            if len(workbook.sheetnames) != 1:
                raise ValueError(
                    "La tabla contiene varias hojas; se debe indicar explícitamente la hoja asociada a la comuna."
                )
            sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = [str(item or "").strip() for item in next(iterator)]
        except StopIteration:
            return [], []
        rows = []
        for values in iterator:
            if not any(value not in (None, "") for value in values):
                continue
            rows.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
        return headers, rows
    return base.read_table(table_path)


def _alias_index(aliases: dict[str, str] | None) -> dict[str, str]:
    return {_key(source): _key(target) for source, target in (aliases or {}).items()}


def _resolved_key(value: Any, aliases: dict[str, str]) -> str:
    key = _key(value)
    return aliases.get(key, key)


def validate_prc_table_pair(
    prc_path: str | Path,
    table_path: str | Path,
    *,
    expected_comuna: str | None = None,
    preferred_layer: str | None = None,
    table_sheet: str | None = None,
    codigo_aliases: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Valida que el PRC y su tabla normativa estén vinculados por CODIGO_PRC.

    El modelo productivo permite:
    - varios polígonos con un mismo CODIGO_PRC;
    - varias filas normativas/variantes para un mismo CODIGO_PRC.

    La cantidad de filas de la tabla se preserva, pero no tiene que igualar la cantidad de
    entidades espaciales. Se bloquea la producción cuando algún código queda sin vínculo.
    """
    prc_path = Path(prc_path)
    table_path = Path(table_path)
    errors: list[str] = []

    if not prc_path.exists():
        errors.append("Falta el archivo PRC.")
    if not table_path.exists():
        errors.append("Falta la tabla normativa asociada.")
    if errors:
        return {
            "valid": False,
            "state": "ERROR_ESTRUCTURAL",
            "errors": errors,
            "prc_path": str(prc_path),
            "table_path": str(table_path),
        }

    try:
        polygon_count, layer, polygon_codes = read_prc_codes(prc_path, preferred_layer)
    except (ValueError, FileNotFoundError) as exc:
        errors.append(str(exc))
        polygon_count, layer, polygon_codes = None, None, []

    try:
        headers, rows = _read_table(table_path, table_sheet)
    except Exception as exc:
        errors.append(f"No se pudo leer la tabla normativa: {exc}")
        headers, rows = [], []

    row_count = len(rows)
    communes = sorted({str(row.get("COMUNA", "")).strip() for row in rows if str(row.get("COMUNA", "")).strip()})
    commune = communes[0] if len(communes) == 1 else ""

    if row_count == 0:
        errors.append("La tabla normativa no contiene filas productivas.")
    if len(communes) > 1:
        errors.append("La tabla contiene más de una comuna y no puede ligarse a un único PRC.")
    if expected_comuna and commune and _key(expected_comuna) != _key(commune):
        errors.append(f"La tabla corresponde a {commune}, no a {expected_comuna}.")

    missing_fields = [field for field in base.FIELDS if field not in headers]
    if missing_fields:
        errors.append("Faltan campos productivos obligatorios: " + ", ".join(missing_fields))

    table_codes = [str(row.get("CODIGO_PRC", "") or "").strip() for row in rows]
    blank_polygon_codes = sum(not code for code in polygon_codes)
    blank_table_codes = sum(not code for code in table_codes)
    if blank_polygon_codes:
        errors.append(f"El PRC contiene {blank_polygon_codes} polígonos sin CODIGO_PRC.")
    if blank_table_codes:
        errors.append(f"La tabla contiene {blank_table_codes} filas sin CODIGO_PRC.")

    aliases = _alias_index(codigo_aliases)
    polygon_keys = {_resolved_key(code, aliases) for code in polygon_codes if code}
    table_keys = {_resolved_key(code, aliases) for code in table_codes if code}
    missing_in_table_keys = polygon_keys - table_keys
    orphan_table_keys = table_keys - polygon_keys

    polygon_raw_by_key: dict[str, set[str]] = {}
    for code in polygon_codes:
        if code:
            polygon_raw_by_key.setdefault(_resolved_key(code, aliases), set()).add(code)
    table_raw_by_key: dict[str, set[str]] = {}
    for code in table_codes:
        if code:
            table_raw_by_key.setdefault(_resolved_key(code, aliases), set()).add(code)

    missing_in_table = sorted({raw for key in missing_in_table_keys for raw in polygon_raw_by_key.get(key, set())})
    orphan_table = sorted({raw for key in orphan_table_keys for raw in table_raw_by_key.get(key, set())})

    if missing_in_table:
        errors.append(
            "Hay CODIGO_PRC presentes en el PRC sin normativa asociada: " + ", ".join(missing_in_table)
        )
    if orphan_table:
        errors.append(
            "Hay filas normativas cuyo CODIGO_PRC no existe en el PRC vigente: " + ", ".join(orphan_table)
        )

    return {
        "valid": not errors,
        "state": "LISTO_AUDITAR" if not errors else "ERROR_VINCULO",
        "errors": errors,
        "comuna": commune or expected_comuna or "",
        "prc_path": str(prc_path),
        "table_path": str(table_path),
        "table_sheet": table_sheet or "",
        "prc_layer": layer,
        "polygon_count": polygon_count,
        "row_count": row_count,
        "distinct_polygon_codes": len({_key(code) for code in polygon_codes if code}),
        "distinct_table_codes": len({_key(code) for code in table_codes if code}),
        "blank_polygon_codes": blank_polygon_codes,
        "blank_table_codes": blank_table_codes,
        "missing_in_table": missing_in_table,
        "orphan_table_codes": orphan_table,
        "productive_fields": len(base.FIELDS),
        "missing_fields": missing_fields,
        "link_field": "CODIGO_PRC",
        "relationship": "muchos polígonos ↔ CODIGO_PRC ↔ una o varias filas normativas",
        "contract": "PRC y tabla obligatorios; todas las filas se preservan; todo CODIGO_PRC debe resolver en ambos lados",
    }
