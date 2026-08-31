from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import tempfile
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from . import engine as base
from . import engine_v2
from .pairing import validate_prc_table_pair


def _key(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


def _safe(value: str) -> str:
    text = unicodedata.normalize("NFD", value).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_").upper()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _load_json(path: str | Path | None, default: Any) -> Any:
    if not path:
        return default
    file_path = Path(path)
    if not file_path.exists():
        return default
    return json.loads(file_path.read_text(encoding="utf-8"))


def _default_state_path() -> Path:
    root = Path(os.environ.get("LOCALAPPDATA") or Path.home() / ".tui_dei")
    return root / "TUI_DEI" / "estado_tablas_normativas.json"


def _master_index(master_path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(master_path, read_only=True, data_only=True)
    index: dict[str, dict[str, Any]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(iterator)]
        except StopIteration:
            continue
        if "COMUNA" not in headers:
            continue
        comuna_index = headers.index("COMUNA")
        first_comuna = ""
        for row in iterator:
            value = row[comuna_index] if comuna_index < len(row) else ""
            if value not in (None, ""):
                first_comuna = str(value).strip()
                break
        if not first_comuna:
            continue
        key = _key(first_comuna)
        if key in index:
            index[key].setdefault("duplicates", []).append(sheet_name)
        else:
            index[key] = {"comuna": first_comuna, "sheet": sheet_name, "duplicates": []}
    return index


def _discover_prc(prc_root: Path) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for folder in prc_root.rglob("*"):
        if not folder.is_dir() or _key(folder.name) != _key("PRC Trabajado"):
            continue
        gpkg_files = sorted(folder.glob("*.gpkg"))
        commune = folder.parent.name
        if len(gpkg_files) == 1:
            results.append({"comuna": commune, "gpkg": gpkg_files[0], "error": ""})
        elif len(gpkg_files) == 0:
            results.append({"comuna": commune, "gpkg": None, "error": "PRC Trabajado no contiene GPKG."})
        else:
            results.append({"comuna": commune, "gpkg": None, "error": "PRC Trabajado contiene más de un GPKG; no se adivina cuál es productivo."})
    return results


def _read_master_sheet(master_path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(master_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(iterator)]
    except StopIteration:
        return [], []
    rows: list[dict[str, Any]] = []
    for values in iterator:
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))})
    return headers, rows


def _aliases_for(catalog: dict[str, Any], comuna: str) -> dict[str, str]:
    registry = catalog.get("por_comuna", {}) if isinstance(catalog, dict) else {}
    target = _key(comuna)
    for name, aliases in registry.items():
        if _key(name) == target:
            return dict(aliases or {})
    return {}


def _coverage_for(catalog: dict[str, Any], comuna: str) -> str:
    registry = catalog.get("por_comuna", {}) if isinstance(catalog, dict) else {}
    target = _key(comuna)
    for name, item in registry.items():
        if _key(name) == target:
            if isinstance(item, str):
                return item.upper()
            return str((item or {}).get("estado", "PENDIENTE")).upper()
    return "PENDIENTE"


def _blocking_findings(findings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    blocked = []
    for finding in findings:
        status = str(finding.get("status", "")).upper()
        confidence = str(finding.get("confidence", "")).upper()
        if status in {"CONFLICTO NORMATIVO", "SIN FUENTE"}:
            blocked.append(finding)
        elif status == "POSIBLE ERROR" and confidence in {"ALTA", "MEDIA"}:
            blocked.append(finding)
    return blocked


def _verify_output(path: Path, expected_rows: int) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = list(next(iterator))
    rows = list(iterator)
    if headers != base.FIELDS:
        raise RuntimeError("La salida no conserva exactamente los 35 campos productivos.")
    if len(rows) != expected_rows:
        raise RuntimeError(
            f"La salida cambió la cantidad de filas normativas: {expected_rows} -> {len(rows)}."
        )


def _atomic_write_normalized(output_path: Path, rows: list[dict[str, Any]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(dir=output_path.parent) as tmp:
        temp_path = Path(tmp) / output_path.name
        base._write_table_xlsx(temp_path, rows)
        _verify_output(temp_path, len(rows))
        os.replace(temp_path, output_path)


def run(
    *,
    prc_root: str | Path,
    master_path: str | Path,
    output_dir: str | Path,
    exact_rules_path: str | Path | None = None,
    conditional_rules_path: str | Path | None = None,
    aliases_path: str | Path | None = None,
    coverage_path: str | Path | None = None,
    state_path: str | Path | None = None,
) -> dict[str, Any]:
    prc_root = Path(prc_root)
    master_path = Path(master_path)
    output_dir = Path(output_dir)
    state_file = Path(state_path) if state_path else _default_state_path()
    state_file.parent.mkdir(parents=True, exist_ok=True)

    previous = _load_json(state_file, {"comunas": {}})
    exact_catalog = base.load_rule_catalog(exact_rules_path)
    conditional_catalog = engine_v2.load_conditional_catalog(conditional_rules_path)
    aliases_catalog = _load_json(aliases_path, {"por_comuna": {}})
    coverage_catalog = _load_json(coverage_path, {"por_comuna": {}})
    master = _master_index(master_path)
    master_hash = _sha256(master_path)

    now = datetime.now(timezone.utc).isoformat()
    result: dict[str, Any] = {
        "schema_version": 1,
        "processed_at": now,
        "master": str(master_path),
        "master_sha256": master_hash,
        "output_dir": str(output_dir),
        "comunas": {},
    }

    for discovered in _discover_prc(prc_root):
        comuna = discovered["comuna"]
        key = _key(comuna)
        item: dict[str, Any] = {
            "comuna": comuna,
            "gpkg": str(discovered["gpkg"] or ""),
            "estado": "PENDIENTE",
            "errores": [],
            "actualizado": now,
        }
        result["comunas"][key] = item

        if discovered["error"]:
            item["estado"] = "ERROR ESTRUCTURAL"
            item["errores"].append(discovered["error"])
            continue
        if key not in master:
            item["estado"] = "FALTA TABLA"
            item["errores"].append("No existe una hoja inequívoca para la comuna en el maestro normativo.")
            continue
        if master[key].get("duplicates"):
            item["estado"] = "ERROR ESTRUCTURAL"
            item["errores"].append(
                "La comuna aparece en más de una hoja del maestro: "
                + ", ".join([master[key]["sheet"], *master[key]["duplicates"]])
            )
            continue

        gpkg_path: Path = discovered["gpkg"]
        sheet = master[key]["sheet"]
        aliases = _aliases_for(aliases_catalog, comuna)
        pair = validate_prc_table_pair(
            gpkg_path,
            master_path,
            expected_comuna=master[key]["comuna"],
            table_sheet=sheet,
            codigo_aliases=aliases,
        )
        item.update({
            "hoja": sheet,
            "poligonos": pair.get("polygon_count"),
            "filas_normativas": pair.get("row_count"),
            "codigos_prc": pair.get("distinct_polygon_codes"),
            "codigos_tabla": pair.get("distinct_table_codes"),
            "sin_normativa": pair.get("missing_in_table", []),
            "sin_geometria": pair.get("orphan_table_codes", []),
        })
        if not pair["valid"]:
            item["estado"] = "ERROR VÍNCULO"
            item["errores"].extend(pair["errors"])
            continue

        coverage = _coverage_for(coverage_catalog, comuna)
        item["cobertura_fuentes"] = coverage
        if coverage != "COMPLETA":
            item["estado"] = "FUENTES INCOMPLETAS"
            item["errores"].append("La comuna aún no tiene catálogo normativo oficial con cobertura completa.")
            continue

        gpkg_hash = _sha256(gpkg_path)
        fingerprint = hashlib.sha256(
            (gpkg_hash + master_hash + sheet + json.dumps(aliases, sort_keys=True, ensure_ascii=False)).encode("utf-8")
        ).hexdigest()
        item["fingerprint"] = fingerprint
        prior = previous.get("comunas", {}).get(key, {})
        output_path = output_dir / f"PRC_{_safe(master[key]['comuna'])}_NORMALIZADO.xlsx"
        if prior.get("fingerprint") == fingerprint and prior.get("estado") == "LISTA PARA STAGING" and output_path.exists():
            item.update(prior)
            item["actualizado"] = now
            item["sin_cambios"] = True
            continue

        headers, rows = _read_master_sheet(master_path, sheet)
        audit = engine_v2.audit_table(headers, rows, exact_catalog, conditional_catalog)
        blocking = _blocking_findings(audit["findings"])
        item.update({
            "correcciones_confirmadas": audit["critical"],
            "posibles": audit["possible"],
            "normalizaciones_formato": audit["formatting"],
            "conflictos": audit["conflicts"],
            "hallazgos_bloqueantes": len(blocking),
            "fingerprint": fingerprint,
        })
        if blocking:
            item["estado"] = "CON OBSERVACIONES"
            item["errores"].extend(
                f"{finding.get('field')}: {finding.get('reason')}" for finding in blocking[:20]
            )
            continue

        if len(audit["rows"]) != len(rows):
            item["estado"] = "ERROR ESTRUCTURAL"
            item["errores"].append("La auditoría alteró la cantidad de filas normativas.")
            continue

        _atomic_write_normalized(output_path, audit["rows"])
        item["estado"] = "LISTA PARA STAGING"
        item["salida"] = str(output_path)
        item["sin_cambios"] = False

    state_file.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Procesa automáticamente pares PRC + tabla normativa de Sistema Operativo DEI.")
    parser.add_argument("--prc-root", required=True)
    parser.add_argument("--master", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--exact-rules", default="config/tablas_normativas_reglas.json")
    parser.add_argument("--conditional-rules", default="config/tablas_normativas_condicionales.json")
    parser.add_argument("--aliases", default="config/tablas_normativas_codigo_aliases.json")
    parser.add_argument("--coverage", default="config/tablas_normativas_cobertura.json")
    parser.add_argument("--state")
    args = parser.parse_args()

    result = run(
        prc_root=args.prc_root,
        master_path=args.master,
        output_dir=args.output,
        exact_rules_path=args.exact_rules,
        conditional_rules_path=args.conditional_rules,
        aliases_path=args.aliases,
        coverage_path=args.coverage,
        state_path=args.state,
    )
    counts: dict[str, int] = {}
    for item in result["comunas"].values():
        counts[item["estado"]] = counts.get(item["estado"], 0) + 1
    print(json.dumps(counts, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
