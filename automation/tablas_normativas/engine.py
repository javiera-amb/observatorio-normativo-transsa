from __future__ import annotations

import csv
import json
import math
import re
import unicodedata
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


FIELDS = [
    "COMUNA", "RIALCOMSII", "CODIGO_PRC", "ZONA", "USO", "SUBZONA_USO", "EDIF", "SUBZONA_EDIF",
    "DEFINICION_ZONA", "ESPECIF_GENERAL", "ESPECIF_ESPECIF", "UPERM", "UPROH", "TABLA",
    "DETALLE_TABLA_ORDENANZA", "DENS_HAB_HA", "DENS_VIV_HA", "SUB_PREDIAL", "CONSTRUCCION",
    "OCUPACION", "OCUPACION_SUP", "PISOS_MAX", "ALTURA_MIN", "ALTURA_MAX", "ARBORIZACION", "PAGE",
    "AREA_LIBRE_MIN", "AGRUPAMIENTO", "RASANTE", "DIST_MEDIANEROS", "ADOSAMIENTO", "ANTEJARDIN",
    "INCENTIVO", "FUENTE", "COMENTS_NORM",
]

NULL_LITERALS = {"NULL", "N/A", "N/D", "-"}
NUMERIC_DECIMAL = {"CONSTRUCCION", "OCUPACION", "OCUPACION_SUP", "ALTURA_MIN", "ALTURA_MAX", "AREA_LIBRE_MIN"}
GROUP_ORDER = ["AISLADO", "PAREADO", "CONTINUO"]
OCR_PATTERN = re.compile(r"(?:^|\s)[OoPp][.,]?\d|\b[lI][.,]?\d", re.I)


@dataclass
class Finding:
    row: int
    field: str
    original: Any
    proposed: Any
    status: str
    confidence: str
    reason: str
    source: str = ""
    page: str = ""
    rule_id: str = ""


def _key(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


def _numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    text = str(value or "").strip()
    if not re.fullmatch(r"-?\d+(?:[.,]\d+)?", text):
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def _same(a: Any, b: Any) -> bool:
    an, bn = _numeric(a), _numeric(b)
    if an is not None and bn is not None:
        return math.isclose(an, bn, rel_tol=0, abs_tol=1e-9)
    return _key(a) == _key(b)


def _row_signature(row: dict[str, Any]) -> str:
    return json.dumps([row.get(field, "") for field in FIELDS], ensure_ascii=False, default=str)


def load_rule_catalog(path: str | Path | None) -> dict[str, Any]:
    if not path:
        return {"exact_rules": []}
    file_path = Path(path)
    if not file_path.exists():
        return {"exact_rules": []}
    return json.loads(file_path.read_text(encoding="utf-8"))


def read_table(path: str | Path) -> tuple[list[str], list[dict[str, Any]]]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = [str(value or "").strip() for value in (reader.fieldnames or [])]
            rows = [{str(k or "").strip(): v for k, v in row.items()} for row in reader]
        return headers, rows
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(iterator)]
        except StopIteration:
            return [], []
        rows = []
        for values in iterator:
            if not any(value not in (None, "") for value in values):
                continue
            rows.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
        return headers, rows
    raise ValueError(f"Formato no soportado: {path.suffix}")


def _write_table_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TABLA_NORMALIZADA"
    sheet.append(FIELDS)
    for row in rows:
        sheet.append([row.get(field, "") for field in FIELDS])
    fill = PatternFill("solid", fgColor="243A5E")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
    sheet.freeze_panes = "A2"
    workbook.save(path)


def _write_qa_xlsx(path: Path, findings: list[dict[str, Any]]) -> None:
    columns = ["row", "field", "original", "proposed", "status", "confidence", "reason", "source", "page", "rule_id"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "QA_TRAZABILIDAD"
    sheet.append(columns)
    for item in findings:
        sheet.append([item.get(column, "") for column in columns])
    fill = PatternFill("solid", fgColor="243A5E")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
    sheet.freeze_panes = "A2"
    workbook.save(path)


def _normalize_text(value: str) -> str:
    value = value.replace("\u00a0", " ")
    value = re.sub(r"[\u200B-\u200D\uFEFF]", "", value)
    return re.sub(r"\s+", " ", value).strip()


def _normalize_grouping(value: Any) -> Any:
    if not isinstance(value, str) or not value.strip():
        return value
    tokens = [_normalize_text(item).upper() for item in re.split(r"[,;/]+", value) if _normalize_text(item)]
    if not tokens or not all(token in GROUP_ORDER for token in tokens):
        return value
    return "; ".join(item for item in GROUP_ORDER if item in tokens)


def _apply_rule(
    row_number: int,
    row: dict[str, Any],
    field: str,
    value: Any,
    findings: list[Finding],
    rule_catalog: dict[str, Any],
) -> Any:
    for rule in rule_catalog.get("exact_rules", []):
        if rule.get("field") != field:
            continue
        if field == "CODIGO_PRC":
            # CODIGO_PRC se preserva siempre en el motor base. Incluso una coincidencia exacta
            # sólo se reporta; cualquier cambio debe pasar por engine_v2 con contexto, fuente y
            # allow_codigo_prc_change explícito.
            if not _same(value, rule.get("original")):
                continue
            findings.append(Finding(
                row_number, field, value, rule.get("corrected", ""), "POSIBLE ERROR",
                str(rule.get("confidence", "MEDIA")),
                str(rule.get("reason", "CODIGO_PRC requiere revisión contextual.")),
                str(rule.get("source", "")), str(rule.get("page", "")), str(rule.get("id", ""))
            ))
            continue
        commune = str(row.get("COMUNA", "") or "")
        rule_commune = str(rule.get("comuna", "") or "")
        if rule_commune and _key(commune) != _key(rule_commune):
            continue
        if not _same(value, rule.get("original")):
            continue
        proposed = rule.get("corrected", "")
        status = str(rule.get("status", "ERROR CONFIRMADO"))
        confidence = str(rule.get("confidence", "ALTA"))
        findings.append(Finding(
            row_number, field, value, proposed, status, confidence,
            str(rule.get("reason", "Regla exacta de fuente oficial.")),
            str(rule.get("source", "")), str(rule.get("page", "")), str(rule.get("id", ""))
        ))
        if rule.get("auto_apply") is True and confidence.upper() == "ALTA" and status == "ERROR CONFIRMADO":
            return proposed
    return value


def _normalize_cell(row_number: int, field: str, value: Any, findings: list[Finding]) -> Any:
    if value is None:
        return ""
    original = value
    if isinstance(value, str):
        proposed = _normalize_text(value)
        if proposed.upper() in NULL_LITERALS:
            findings.append(Finding(
                row_number, field, original, "", "NORMALIZACIÓN DE FORMATO", "ALTA",
                "Literal de ausencia reemplazado por celda vacía."
            ))
            return ""
        if proposed != value:
            findings.append(Finding(
                row_number, field, original, proposed, "NORMALIZACIÓN DE FORMATO", "ALTA",
                "Espacios o caracteres invisibles normalizados."
            ))
        value = proposed

    if field == "AGRUPAMIENTO":
        proposed = _normalize_grouping(value)
        if proposed != value:
            findings.append(Finding(
                row_number, field, original, proposed, "NORMALIZACIÓN DE FORMATO", "ALTA",
                "Agrupamientos normalizados en una sola celda; no se crean filas adicionales."
            ))
            value = proposed

    if field in NUMERIC_DECIMAL and isinstance(value, str):
        num = _numeric(value)
        if num is not None:
            if str(value) != str(num):
                findings.append(Finding(
                    row_number, field, original, num, "NORMALIZACIÓN DE FORMATO", "ALTA",
                    "Valor convertido a tipo numérico sin alterar su magnitud."
                ))
            value = num

    if field == "PISOS_MAX" and isinstance(value, str):
        num = _numeric(value)
        if num is not None and num.is_integer():
            proposed = int(num)
            if str(value) != str(proposed):
                findings.append(Finding(
                    row_number, field, original, proposed, "NORMALIZACIÓN DE FORMATO", "ALTA",
                    "PISOS_MAX convertido a entero."
                ))
            value = proposed

    return value


def audit_table(
    headers: list[str], rows: list[dict[str, Any]], rule_catalog: dict[str, Any] | None = None
) -> dict[str, Any]:
    rule_catalog = rule_catalog or {"exact_rules": []}
    findings: list[Finding] = []
    normalized_rows: list[dict[str, Any]] = []
    missing = [field for field in FIELDS if field not in headers]
    extras = [field for field in headers if field not in FIELDS]

    if missing:
        findings.append(Finding(
            0, "ESTRUCTURA", ", ".join(missing), "", "ERROR CONFIRMADO", "ALTA",
            "Faltan campos productivos obligatorios."
        ))
    if extras:
        findings.append(Finding(
            0, "ESTRUCTURA", ", ".join(extras), "", "POSIBLE ERROR", "MEDIA",
            "Existen columnas adicionales; se excluyen de la salida de 35 campos sin alterar las filas."
        ))
    if len(headers) >= len(FIELDS) and not missing and headers[:len(FIELDS)] != FIELDS:
        findings.append(Finding(
            0, "ESTRUCTURA", " | ".join(headers[:len(FIELDS)]), " | ".join(FIELDS),
            "NORMALIZACIÓN DE FORMATO", "ALTA", "Los campos productivos se ordenan al contrato oficial de 35 campos."
        ))

    # Regla estructural central: cada fila normativa de entrada produce exactamente una fila de salida.
    for index, source_row in enumerate(rows, start=2):
        row = {field: _normalize_cell(index, field, source_row.get(field, ""), findings) for field in FIELDS}
        for field in FIELDS:
            row[field] = _apply_rule(index, row, field, row[field], findings, rule_catalog)
        normalized_rows.append(row)

        code = str(row.get("CODIGO_PRC", "") or "").strip()
        if code and (len(code) > 45 or len(re.findall(r"\s+", code)) >= 4):
            findings.append(Finding(
                index, "CODIGO_PRC", code, "", "POSIBLE ERROR", "MEDIA",
                "CODIGO_PRC parece descriptivo; se conserva hasta demostrar que el identificador es incorrecto."
            ))

        pisos = row.get("PISOS_MAX")
        pisos_num = _numeric(pisos)
        if pisos_num is not None and not pisos_num.is_integer():
            findings.append(Finding(
                index, "PISOS_MAX", pisos, "", "POSIBLE ERROR", "ALTA",
                "PISOS_MAX es decimal; no se trunca sin fuente oficial."
            ))

        for field in ["CONSTRUCCION", "OCUPACION", "OCUPACION_SUP", "ALTURA_MIN", "ALTURA_MAX", "AREA_LIBRE_MIN"]:
            value = row.get(field)
            num = _numeric(value)
            if num is not None and num < 0:
                findings.append(Finding(
                    index, field, value, "", "POSIBLE ERROR", "MEDIA",
                    "Valor negativo que normalmente no corresponde a este parámetro."
                ))
            if isinstance(value, str) and OCR_PATTERN.search(value):
                findings.append(Finding(
                    index, field, value, "", "POSIBLE ERROR", "MEDIA",
                    "Patrón compatible con OCR O/0, P/0 o I/l/1; no se autocorrige sin fuente."
                ))

        amin = _numeric(row.get("ALTURA_MIN"))
        amax = _numeric(row.get("ALTURA_MAX"))
        if amin is not None and amax is not None and amin > amax:
            findings.append(Finding(
                index, "ALTURA_MIN / ALTURA_MAX", f"{row.get('ALTURA_MIN')} > {row.get('ALTURA_MAX')}", "",
                "ERROR CONFIRMADO", "ALTA", "ALTURA_MIN es mayor que ALTURA_MAX."
            ))

        if re.search(r"AISLADO|PAREADO|CONTINUO", str(row.get("AREA_LIBRE_MIN", "")), re.I):
            findings.append(Finding(
                index, "AREA_LIBRE_MIN", row.get("AREA_LIBRE_MIN"), "", "POSIBLE ERROR", "MEDIA",
                "Contenido compatible con AGRUPAMIENTO; posible desplazamiento de columnas."
            ))
        if re.search(r"AISLADO|PAREADO|CONTINUO", str(row.get("RASANTE", "")), re.I):
            findings.append(Finding(
                index, "RASANTE", row.get("RASANTE"), "", "POSIBLE ERROR", "MEDIA",
                "Contenido incompatible con rasante; posible desplazamiento de columnas."
            ))

    # Una fila repetida puede ser una variante válida o una repetición técnica. Nunca se elimina.
    # El vínculo con geometría se valida por CODIGO_PRC, no por igualdad de filas.
    seen: dict[str, int] = {}
    for index, row in enumerate(normalized_rows, start=2):
        signature = _row_signature(row)
        if signature in seen:
            findings.append(Finding(
                index, "FILA", f"Campos productivos iguales a fila {seen[signature]}", "", "POSIBLE ERROR", "BAJA",
                "Se conserva la fila: puede corresponder a una variante válida o a una repetición técnica. No se fusiona ni elimina automáticamente."
            ))
        else:
            seen[signature] = index

    assert len(normalized_rows) == len(rows), "La auditoría no puede cambiar la cantidad de filas normativas."

    critical = sum(f.status == "ERROR CONFIRMADO" for f in findings)
    possible = sum(f.status == "POSIBLE ERROR" for f in findings)
    formatting = sum(f.status == "NORMALIZACIÓN DE FORMATO" for f in findings)
    conflicts = sum(f.status == "CONFLICTO NORMATIVO" for f in findings)

    return {
        "headers": headers,
        "rows": normalized_rows,
        "findings": [asdict(item) for item in findings],
        "missing": missing,
        "extras": extras,
        "critical": critical,
        "possible": possible,
        "formatting": formatting,
        "conflicts": conflicts,
        "input_rows": len(rows),
        "output_rows": len(normalized_rows),
        "processed_at": datetime.now(timezone.utc).isoformat(),
    }


def process_file(
    input_path: str | Path,
    normalized_dir: str | Path,
    qa_dir: str | Path,
    rule_catalog_path: str | Path | None = None,
) -> dict[str, Any]:
    input_path = Path(input_path)
    headers, rows = read_table(input_path)
    catalog = load_rule_catalog(rule_catalog_path)
    result = audit_table(headers, rows, catalog)
    commune = str(result["rows"][0].get("COMUNA", "SIN_COMUNA") if result["rows"] else "SIN_COMUNA")
    safe_commune = re.sub(r"[^A-Za-z0-9ÁÉÍÓÚÜÑáéíóúüñ]+", "_", commune).strip("_").upper()

    normalized_dir = Path(normalized_dir)
    qa_dir = Path(qa_dir)
    normalized_dir.mkdir(parents=True, exist_ok=True)
    qa_dir.mkdir(parents=True, exist_ok=True)

    normalized_path = normalized_dir / f"PRC_{safe_commune}_NORMALIZADO.xlsx"
    qa_path = qa_dir / f"QA_PRC_{safe_commune}.xlsx"
    status_path = qa_dir / f"STATUS_PRC_{safe_commune}.json"

    _write_table_xlsx(normalized_path, result["rows"])
    _write_qa_xlsx(qa_path, result["findings"])
    status = {
        "comuna": commune,
        "source": input_path.name,
        "normalized": normalized_path.name,
        "qa": qa_path.name,
        "rows": result["input_rows"],
        "fields": len(FIELDS),
        "findings": len(result["findings"]),
        "errores_confirmados": result["critical"],
        "posibles_errores": result["possible"],
        "normalizaciones_formato": result["formatting"],
        "conflictos_normativos": result["conflicts"],
        "state": "CON OBSERVACIONES" if result["critical"] or result["possible"] or result["conflicts"] else "VALIDADA",
        "processed_at": result["processed_at"],
    }
    status_path.write_text(json.dumps(status, ensure_ascii=False, indent=2), encoding="utf-8")
    return {**result, "normalized_path": normalized_path, "qa_path": qa_path, "status_path": status_path, "status": status}
