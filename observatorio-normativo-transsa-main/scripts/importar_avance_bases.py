#!/usr/bin/env python3
"""Importa el inventario inicial de producción sin usar SharePoint.

El Excel es una carga inicial. Después de importarlo, el estado compartido se
versiona en Git y los archivos SIG continúan almacenados en OneDrive.
"""

from __future__ import annotations

import argparse
import json
import unicodedata
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ESTADOS = {
    "pendiente": "pendiente",
    "en proceso": "en_desarrollo",
    "en desarrollo": "en_desarrollo",
    "actualizado": "actualizado",
    "actualizada": "actualizado",
    "enviado": "enviado",
    "enviada": "enviado",
}

QAS = {
    "sin qa": "pendiente",
    "pendiente": "pendiente",
    "observado": "observaciones",
    "observaciones": "observaciones",
    "aprobado": "aprobado",
    "aprobada": "aprobado",
}

CAPAS = {
    "barrios": ("BARRIOS", "QA BARRIOS", "Encargado2"),
    "areas_homogeneas_2022": ("AH 2022", "QA AH 2022", "Encargado3"),
    "predios": ("PREDIOS", "QA PREDIOS", "Encargado4"),
    "poligonos_permisos": ("POLIGONOS PERMISOS", "QA POLIGONOS PERMISOS", "Encargado5"),
}


def texto(valor: Any) -> str:
    return str(valor or "").strip()


def normalizar(valor: Any) -> str:
    salida = unicodedata.normalize("NFD", texto(valor).lower())
    return "".join(caracter for caracter in salida if unicodedata.category(caracter) != "Mn")


def estado(valor: Any) -> str:
    return ESTADOS.get(normalizar(valor), "pendiente")


def qa(valor: Any) -> str:
    return QAS.get(normalizar(valor), "pendiente")


def registro_capa(fila: dict[str, Any], columnas: tuple[str, str, str]) -> dict[str, Any]:
    campo_estado, campo_qa, campo_responsable = columnas
    return {
        "estado_equipo": estado(fila.get(campo_estado)),
        "qa_interno": qa(fila.get(campo_qa)),
        "responsable": texto(fila.get(campo_responsable)) or None,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa Avance_Bases_de_datos.xlsx al portal TUI.")
    parser.add_argument("excel", type=Path)
    parser.add_argument("--salida", type=Path, default=Path("data/avance_bases_inicial.js"))
    parser.add_argument("--hoja", default="Comunas de Chile")
    args = parser.parse_args()

    libro = load_workbook(args.excel, read_only=True, data_only=True)
    if args.hoja not in libro.sheetnames:
        raise SystemExit(f"No existe la hoja {args.hoja!r}. Hojas: {', '.join(libro.sheetnames)}")
    hoja = libro[args.hoja]
    filas = hoja.iter_rows(values_only=True)
    encabezados = [texto(valor) for valor in next(filas)]
    requeridos = {"Región", "Comuna", "Código SII", "PRC", "QA PRC", "Encargado"}
    faltantes = sorted(requeridos - set(encabezados))
    if faltantes:
        raise SystemExit(f"Faltan columnas obligatorias: {', '.join(faltantes)}")

    comunas: dict[str, Any] = {}
    conteo_prc: Counter[str] = Counter()
    conteos_capas: dict[str, Counter[str]] = {nombre: Counter() for nombre in CAPAS}
    for valores in filas:
        fila = dict(zip(encabezados, valores))
        region = texto(fila.get("Región"))
        comuna = texto(fila.get("Comuna"))
        if not region or not comuna:
            continue
        estado_prc = estado(fila.get("PRC"))
        conteo_prc[estado_prc] += 1
        capas = {nombre: registro_capa(fila, columnas) for nombre, columnas in CAPAS.items()}
        for nombre, registro in capas.items():
            conteos_capas[nombre][registro["estado_equipo"]] += 1
        clave = f"{region}|{comuna}"
        comunas[clave] = {
            "region": region,
            "comuna": comuna,
            "codigo_sii": texto(fila.get("Código SII")).zfill(5),
            "prc": {
                "estado_produccion": estado_prc,
                "qa_revision_javiera": qa(fila.get("QA PRC")),
                "responsable": texto(fila.get("Encargado")) or None,
                "alerta_sin_modificaciones": texto(fila.get("SIN MODIFICACIONES")) or None,
            },
            "capas": capas,
        }

    if len(comunas) != 346:
        raise SystemExit(f"Se esperaban 346 comunas y se importaron {len(comunas)}")

    salida = {
        "schema_version": 1,
        "importado_en": date.today().isoformat(),
        "fuente": args.excel.name,
        "nota": "Carga inicial versionada. No acredita cobertura territorial ni reemplaza el cruce geométrico.",
        "resumen": {
            "comunas": len(comunas),
            "prc": dict(sorted(conteo_prc.items())),
            "capas": {nombre: dict(sorted(conteo.items())) for nombre, conteo in conteos_capas.items()},
        },
        "comunas": comunas,
    }
    args.salida.parent.mkdir(parents=True, exist_ok=True)
    args.salida.write_text(
        "window.AVANCE_BASES_DATOS = " + json.dumps(salida, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(
        f"Importación generada: {args.salida} · comunas={len(comunas)} · "
        f"enviados={conteo_prc['enviado']} · en_desarrollo={conteo_prc['en_desarrollo']} · "
        f"actualizados={conteo_prc['actualizado']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
